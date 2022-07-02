import sys

import openpyxl
import os
import time


"""
根据列名找列
"""
def findcolnumbyname(name,ws):
    for clnum in range(1, ws.max_column + 1):
        if name in ws.cell(row=1, column=clnum).value:
            return clnum

if __name__ == '__main__':
    filename = sys.argv[1]
    filepath = os.path.join(os.path.abspath("."), filename)
    workbook = openpyxl.load_workbook(filepath)
    ws = workbook.worksheets[0]
    data_count = findcolnumbyname("数据(data)", ws)

    maxcolum = ws.max_column + 1
    with open('data.txt','w') as f:
        for i in range(2, ws.max_row + 1):
            datas = ws.cell(row=i, column=data_count).value
            if datas is not None:
                f.write(datas+"\n")

    print("从excel里面提取datas成功，并记录在data.txt文件中")



